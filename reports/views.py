from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

from timetable.models import Timetable, TimeSlot, TimetableEntry
from faculty.models import Faculty
from departments.models import Department, AcademicYear, Section


DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']


@login_required
def reports_dashboard(request):
    context = {
        'timetables': Timetable.objects.select_related('section__department', 'academic_year').filter(is_active=True)[:10],
        'faculty_list': Faculty.objects.select_related('user', 'department').filter(is_active=True)[:10],
        'departments': Department.objects.all(),
        'academic_years': AcademicYear.objects.all(),
    }
    return render(request, 'reports/dashboard.html', context)


# ─── Timetable PDF ───────────────────────────────────────────────────────────

@login_required
def timetable_pdf(request, pk):
    timetable = get_object_or_404(Timetable, pk=pk)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=14, alignment=TA_CENTER, spaceAfter=6)
    subtitle_style = ParagraphStyle('subtitle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, spaceAfter=12)

    story = []
    story.append(Paragraph("CLASS TIMETABLE", title_style))
    story.append(Paragraph(
        f"Section: {timetable.section} | Academic Year: {timetable.academic_year}",
        subtitle_style
    ))

    # Build grid
    time_slots = TimeSlot.objects.filter(is_break=False).order_by('period_number').distinct('period_number')
    entries = {
        (e.time_slot.day, e.time_slot.period_number): e
        for e in timetable.entries.select_related('time_slot', 'subject', 'faculty__user', 'classroom')
    }

    # Header row
    header = ['Period / Day'] + DAYS
    data = [header]

    all_slots = TimeSlot.objects.order_by('period_number', 'day')
    periods = sorted(set(s.period_number for s in all_slots))

    for period in periods:
        slot_obj = TimeSlot.objects.filter(period_number=period).first()
        if slot_obj:
            row_label = f"P{period}\n{slot_obj.start_time.strftime('%I:%M')}-{slot_obj.end_time.strftime('%I:%M')}"
        else:
            row_label = f"P{period}"
        row = [row_label]
        for day in DAYS:
            entry = entries.get((day, period))
            if entry:
                if entry.is_free:
                    row.append("FREE")
                else:
                    cell = f"{entry.subject.code if entry.subject else ''}\n"
                    cell += f"{entry.faculty.full_name[:15] if entry.faculty else ''}\n"
                    cell += f"{entry.classroom.room_number if entry.classroom else ''}"
                    row.append(cell)
            else:
                row.append("-")
        data.append(row)

    col_widths = [3*cm] + [3.5*cm] * len(DAYS)
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e8eaf6')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#283593')),
        ('TEXTCOLOR', (0, 1), (0, -1), colors.white),
    ]))
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="timetable_{timetable.section.department.code}_{timetable.academic_year.year_label}.pdf"'
    return response


# ─── Faculty Timetable PDF ───────────────────────────────────────────────────

@login_required
def faculty_timetable_pdf(request, pk):
    faculty = get_object_or_404(Faculty, pk=pk)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=14, alignment=TA_CENTER)
    story = []
    story.append(Paragraph(f"FACULTY TIMETABLE: {faculty.full_name.upper()}", title_style))
    story.append(Spacer(1, 0.5*cm))

    entries = faculty.timetable_entries.select_related(
        'time_slot', 'subject', 'classroom', 'timetable__section'
    ).filter(is_free=False).order_by('time_slot__day', 'time_slot__period_number')

    all_slots = TimeSlot.objects.order_by('period_number').distinct('period_number')
    periods = sorted(set(s.period_number for s in TimeSlot.objects.all()))
    entry_map = {(e.time_slot.day, e.time_slot.period_number): e for e in entries}

    header = ['Period'] + DAYS
    data = [header]
    for period in periods:
        slot = TimeSlot.objects.filter(period_number=period).first()
        row = [f"P{period}\n{slot.start_time.strftime('%H:%M') if slot else ''}"]
        for day in DAYS:
            e = entry_map.get((day, period))
            if e:
                row.append(f"{e.subject.code}\n{e.timetable.section}\n{e.classroom.room_number if e.classroom else ''}")
            else:
                row.append("-")
        data.append(row)

    table = Table(data, colWidths=[2.5*cm] + [3.5*cm]*len(DAYS), repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#004d40')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e0f2f1')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="faculty_timetable_{faculty.employee_id}.pdf"'
    return response


# ─── Excel Export ────────────────────────────────────────────────────────────

@login_required
def timetable_excel(request, pk):
    timetable = get_object_or_404(Timetable, pk=pk)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f"CLASS TIMETABLE | {timetable.section} | {timetable.academic_year}"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = center_align
    title_cell.fill = PatternFill(start_color="283593", end_color="283593", fill_type="solid")
    title_cell.font = Font(bold=True, color="FFFFFF", size=13)

    # Header row
    headers = ['Period / Time'] + DAYS
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    entries = {
        (e.time_slot.day, e.time_slot.period_number): e
        for e in timetable.entries.select_related('time_slot', 'subject', 'faculty__user', 'classroom')
    }
    periods = sorted(set(
        s.period_number for s in TimeSlot.objects.all()
    ))

    alt_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
    row_num = 3
    for period in periods:
        slot = TimeSlot.objects.filter(period_number=period).first()
        label = f"P{period}\n{slot.start_time.strftime('%H:%M')}-{slot.end_time.strftime('%H:%M')}" if slot else f"P{period}"
        fill = alt_fill if (row_num % 2 == 0) else PatternFill(fill_type=None)

        cell = ws.cell(row=row_num, column=1, value=label)
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = border
        cell.fill = PatternFill(start_color="283593", end_color="283593", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

        for col, day in enumerate(DAYS, 2):
            entry = entries.get((day, period))
            if entry:
                if entry.is_free:
                    val = "FREE"
                else:
                    val = (
                        f"{entry.subject.code if entry.subject else ''}\n"
                        f"{entry.faculty.full_name if entry.faculty else ''}\n"
                        f"{entry.classroom.room_number if entry.classroom else ''}"
                    )
            else:
                val = "-"
            c = ws.cell(row=row_num, column=col, value=val)
            c.alignment = center_align
            c.border = border
            c.fill = fill
        ws.row_dimensions[row_num].height = 50
        row_num += 1

    # Column widths
    ws.column_dimensions['A'].width = 18
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="timetable_{timetable.pk}.xlsx"'
    return response


# ─── Classroom Utilization Report ────────────────────────────────────────────

@login_required
def classroom_utilization(request):
    from classrooms.models import Classroom
    from django.db.models import Count
    classrooms = Classroom.objects.annotate(
        entry_count=Count('timetable_entries')
    ).order_by('-entry_count')
    return render(request, 'reports/classroom_utilization.html', {'classrooms': classrooms})


# ─── Department Report ───────────────────────────────────────────────────────

@login_required
def department_report(request):
    departments = Department.objects.prefetch_related(
        'sections__timetables',
        'faculty_members__user',
        'subjects',
    ).all()
    return render(request, 'reports/department_report.html', {'departments': departments})
